"""Read draft state back from a previously-exported XLSX.

Companion to ``excel_writer.py``. The workbook is the single source of
truth: six ``veryHidden`` sheets under the ``_data_*`` prefix carry the
user's working state as per-entity flat tables plus a JSON-chunked
solver-results cache. See ``_write_data_sheets`` for the writer side.

Two public entry points:

  read_draft_state(file)
      Returns the parsed dict, or raises one of the typed exceptions
      below. UI code maps each exception to a specific user message.
      Structural failures ("not our file", "schema too new") still
      raise — there's nothing meaningful to partition at cell level.

  validate_against_local_data(state, catalog_ids, prof_ids, room_ids)
      Filters offerings/locks against the locally-loaded reference data
      and returns (cleaned_state, errors). Each dropped record emits
      one structured error ``{sheet, row, column, reason, severity}``.
      Partition, don't throw — never refuse the whole file because of
      one bad ref. The Data Issues panel (Phase 3) renders these as a
      clickable list so users can fix offending rows inline.
"""
from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path
from typing import BinaryIO, Union

import openpyxl

from export.excel_writer import (
    DATA_MARKER,
    DATA_SCHEMA_VERSION,
    DATA_SHEET_LOCKED,
    DATA_SHEET_META,
    DATA_SHEET_OFFERINGS,
    DATA_SHEET_PROFESSORS,
    DATA_SHEET_ROOMS,
    DATA_SHEET_SOLVER_RESULTS,
    DATA_SHEET_TUNED_WEIGHTS,
)


# ---------------------------------------------------------------------------
# Exceptions — UI code catches StateReadError to handle any reader failure,
# or specific subclasses to give targeted error copy.
# ---------------------------------------------------------------------------

class StateReadError(Exception):
    """Base class for all reader failures."""


class MissingStateSheet(StateReadError):
    """No _data_meta sheet — not a v2-format Scheduler export."""


class MarkerMismatch(StateReadError):
    """_data_meta's marker row doesn't carry our signature."""


class SchemaVersionUnsupported(StateReadError):
    """schema_version is newer than what this build knows how to read."""

    def __init__(self, found: int, supported: int) -> None:
        super().__init__(f"state schema v{found} > supported v{supported}")
        self.found = found
        self.supported = supported


class MalformedState(StateReadError):
    """Required sheet structure missing, or JSON parse failed."""


# ---------------------------------------------------------------------------
# Sheet readers
# ---------------------------------------------------------------------------

# Required keys in the reconstructed state dict. offerings is also required
# but handled separately — a missing sheet becomes an empty list.
_REQUIRED_KEYS = frozenset({
    "schema_version", "quarter", "year", "solver_mode",
})


def _maybe_parse_json(v):
    """If ``v`` looks like a JSON array/object, try to parse it; fall back
    to the original value on any failure. Used both for kv value rows and
    flat-table cells so nested structures survive a roundtrip."""
    if not isinstance(v, str) or not v:
        return v
    if v[0] not in "[{":
        return v
    try:
        return json.loads(v)
    except (json.JSONDecodeError, TypeError):
        return v


def _read_kv_sheet(ws) -> dict:
    """Two-column key/value sheet → dict. Stops at first empty-key row."""
    out: dict = {}
    for row in ws.iter_rows(values_only=True):
        if not row:
            continue
        key = row[0]
        if key is None or key == "":
            break
        value = row[1] if len(row) > 1 else None
        out[str(key)] = _maybe_parse_json(value)
    return out


def _read_flat_table_sheet(ws) -> list[dict]:
    """Row-per-entity flat table → list[dict]. First row = headers; empty
    rows (all-None) terminate the table.

    Empty cells are treated as absent keys (not ``None``) — the writer's
    union-of-keys layout forces every row to carry every column, so we
    need an unambiguous "this entity doesn't have this field" encoding.
    Our state model treats explicit ``None`` and absent-key identically,
    so this lossless for our shapes."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers: list[str] = [h for h in rows[0] if h is not None]
    if not headers:
        return []
    out: list[dict] = []
    for data_row in rows[1:]:
        if data_row is None or all(v is None for v in data_row):
            continue
        entity: dict = {}
        for i, h in enumerate(headers):
            value = data_row[i] if i < len(data_row) else None
            if value is None:
                continue
            entity[h] = _maybe_parse_json(value)
        out.append(entity)
    return out


def _read_json_chunks_sheet(ws):
    """Column-A JSON chunks → parsed payload. Returns ``None`` if empty."""
    chunks: list[str] = []
    for row in ws.iter_rows(values_only=True):
        if not row:
            break
        v = row[0]
        if v is None or v == "":
            break
        chunks.append(str(v))
    if not chunks:
        return None
    try:
        return json.loads("".join(chunks))
    except json.JSONDecodeError as e:
        raise MalformedState(f"JSON parse failed in chunked sheet: {e}") from e


# ---------------------------------------------------------------------------
# Reader
# ---------------------------------------------------------------------------

def read_draft_state(file: Union[str, Path, BinaryIO, bytes]) -> dict:
    """Open an XLSX and return the embedded draft state dict.

    ``file`` may be a path, a file-like object (Streamlit's UploadedFile
    qualifies), or raw bytes.

    Raises one of the typed exceptions above on any failure.
    """
    if isinstance(file, bytes):
        wb = openpyxl.load_workbook(BytesIO(file), read_only=True, data_only=True)
    else:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)

    if DATA_SHEET_META not in wb.sheetnames:
        raise MissingStateSheet(
            "No _data_meta sheet — this file wasn't exported by this version "
            "of the Scheduler, or isn't a Scheduler export."
        )

    meta = _read_kv_sheet(wb[DATA_SHEET_META])

    marker = meta.pop("marker", None)
    if marker != DATA_MARKER:
        raise MarkerMismatch(f"expected marker {DATA_MARKER!r}, got {marker!r}")

    schema = meta.pop("schema_version", None)
    if not isinstance(schema, int):
        raise MalformedState(
            f"schema_version must be int, got {type(schema).__name__}"
        )
    if schema > DATA_SCHEMA_VERSION:
        raise SchemaVersionUnsupported(schema, DATA_SCHEMA_VERSION)

    # Remaining meta keys carry into the public dict verbatim
    state: dict = dict(meta)
    state["schema_version"] = schema

    # List-valued sheets — absent sheets mean the key wasn't written
    for key, sheet_name in [
        ("offerings",          DATA_SHEET_OFFERINGS),
        ("professors",         DATA_SHEET_PROFESSORS),
        ("rooms",              DATA_SHEET_ROOMS),
        ("locked_assignments", DATA_SHEET_LOCKED),
    ]:
        if sheet_name in wb.sheetnames:
            state[key] = _read_flat_table_sheet(wb[sheet_name])

    # Optional scalar sheet
    if DATA_SHEET_TUNED_WEIGHTS in wb.sheetnames:
        state["tunedWeights"] = _read_kv_sheet(wb[DATA_SHEET_TUNED_WEIGHTS])

    # Optional JSON-chunked cache
    if DATA_SHEET_SOLVER_RESULTS in wb.sheetnames:
        sr = _read_json_chunks_sheet(wb[DATA_SHEET_SOLVER_RESULTS])
        if sr is not None:
            state["solver_results"] = sr

    missing = _REQUIRED_KEYS - state.keys()
    if missing:
        raise MalformedState(f"missing required keys: {sorted(missing)}")

    return state


# ---------------------------------------------------------------------------
# Reference-data drift validation
# ---------------------------------------------------------------------------

def validate_against_local_data(
    state: dict,
    catalog_ids: set[str],
    prof_ids: set[str],
    room_ids: set[str],
) -> tuple[dict, list[dict]]:
    """Filter draft state against local reference data.

    Returns ``(cleaned_state, errors)``. Offerings whose ``catalog_id``
    isn't in ``catalog_ids`` are dropped. Locks whose ``prof_id`` or
    ``room_id`` aren't local are dropped. Each dropped record emits one
    structured error::

        {"sheet": "_data_offerings",   # source _data_* sheet name
         "row": 4,                      # 1-based sheet row (1 = header)
         "column": "catalog_id",        # header field name
         "reason": "Unknown …",         # user-facing detail
         "severity": "error"}           # "error" | "warning" | "info"

    ``row`` is derived from the list index in ``state`` — record N in the
    parsed list lives at sheet row N+2 (+1 for the header row, +1 for
    1-based indexing). ``column`` is the header name rather than an
    Excel letter since the union-of-keys writer doesn't pin a column to
    a field, and the Data Issues panel renders friendly names anyway.

    The cleaned state is a shallow copy with ``offerings`` and
    ``locked_assignments`` replaced; all other keys pass through.
    """
    errors: list[dict] = []

    raw_offerings = state.get("offerings", []) or []
    valid_offerings: list[dict] = []
    for idx, o in enumerate(raw_offerings):
        cid = o.get("catalog_id")
        if cid in catalog_ids:
            valid_offerings.append(o)
        else:
            errors.append({
                "sheet":    DATA_SHEET_OFFERINGS,
                "row":      idx + 2,
                "column":   "catalog_id",
                "reason":   (
                    f"Unknown catalog_id {cid!r} — not in your local catalog"
                    if cid else "Missing catalog_id"
                ),
                "severity": "error",
            })

    raw_locks = state.get("locked_assignments", []) or []
    valid_locks: list[dict] = []
    for idx, la in enumerate(raw_locks):
        prof_id = la.get("prof_id")
        room_id = la.get("room_id")
        if prof_id not in prof_ids:
            errors.append({
                "sheet":    DATA_SHEET_LOCKED,
                "row":      idx + 2,
                "column":   "prof_id",
                "reason":   f"Unknown prof_id {prof_id!r} — not on your roster",
                "severity": "error",
            })
            continue
        if room_id not in room_ids:
            errors.append({
                "sheet":    DATA_SHEET_LOCKED,
                "row":      idx + 2,
                "column":   "room_id",
                "reason":   f"Unknown room_id {room_id!r} — not in your local rooms",
                "severity": "error",
            })
            continue
        valid_locks.append(la)

    cleaned = dict(state)
    cleaned["offerings"] = valid_offerings
    cleaned["locked_assignments"] = valid_locks
    return cleaned, errors
