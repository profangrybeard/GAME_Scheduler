"""Write a formatted 4-sheet Excel workbook from scheduler results.

Usage:
    from export.excel_writer import write_excel
    path = write_excel(results, output_dir)

Sheets
------
  1. Summary       — mode comparison table, quarter overview
  2. Affinity First — schedule grid for affinity_first mode
  3. Time Pref First — schedule grid for time_pref_first mode
  4. Balanced       — schedule grid for balanced mode

Visual conventions
------------------
  Row background  — department colour (game=blue, motion_media=purple, ai=green)
  Yellow cell     — professor is in preferred or override list (good affinity, levels 0-1)
  Orange cell     — time slot is 'not_preferred' for this professor
  Bold header rows
"""

from pathlib import Path
from datetime import date, datetime
import json

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.protection import WorkbookProtection

from config import TIME_SLOTS, DAY_GROUPS


# ---------------------------------------------------------------------------
# Disaster-recovery backups
# ---------------------------------------------------------------------------
# Every xlsx written through write_excel() also lands a timestamped copy in
# <backup_root>/.backups/. Keeps the BACKUP_RETENTION newest; older ones
# prune automatically. Opt-in per-call: pass `backup_root=` to enable.
# Hosted container's filesystem is ephemeral, so backups there are a no-op
# in practice — only the local stack actually benefits.

BACKUPS_DIRNAME  = ".backups"
BACKUP_RETENTION = 10


# ---------------------------------------------------------------------------
# Hidden data-sheet protocol — see _write_data_sheets for details.
# ---------------------------------------------------------------------------
# The workbook is the single source of truth for the user's working state.
# Six veryHidden sheets under the _data_* prefix split that state into
# structured, inspectable tables rather than one opaque JSON blob.

DATA_SCHEMA_VERSION = 2
DATA_MARKER         = "GAME_SCHEDULER_DATA_V2"

DATA_SHEET_META           = "_data_meta"
DATA_SHEET_OFFERINGS      = "_data_offerings"
DATA_SHEET_PROFESSORS     = "_data_professors"
DATA_SHEET_ROOMS          = "_data_rooms"
DATA_SHEET_TUNED_WEIGHTS  = "_data_tuned_weights"
DATA_SHEET_SOLVER_RESULTS = "_data_solver_results"
DATA_SHEET_LOCKED         = "_data_locked_assignments"

# draft_state top-level keys that map to row-per-entity flat-table sheets.
_DATA_LIST_SHEETS: dict[str, str] = {
    "offerings":           DATA_SHEET_OFFERINGS,
    "professors":          DATA_SHEET_PROFESSORS,
    "rooms":               DATA_SHEET_ROOMS,
    "locked_assignments":  DATA_SHEET_LOCKED,
}

# Excel's hard cell limit is 32,767 chars; chunk well under that.
_DATA_CHUNK_SIZE = 30_000

# Cosmetic structure protection — prevents chairs from unhiding sheets via
# Excel's UI. The password lives in this open-source file; it's a "please
# do not edit" barrier, not security against a motivated user.
_DATA_STRUCTURE_PASSWORD = "GAME_SCHEDULER_PROTECTED"

_COMMENT_AUTHOR     = "GAME Scheduler"

# Hover-tooltip copy for the Summary sheet's mode comparison columns.
# Keys are 1-based column numbers matching the _header_row layout.
_SUMMARY_COMMENTS = {
    3: ("Penalty Score",
        "Total penalty for soft preferences missed: professor-course "
        "affinity, time-of-day fit, day-of-week balance. Lower is better. "
        "Hard rules (no double-booking, room capacity) are enforced "
        "absolutely and don't add to this score."),
    4: ("Placed",
        "How many class sections the solver fit into a time slot, out of "
        "all it tried. Anything less than full means some sections couldn't "
        "fit — see Unscheduled."),
    5: ("Unscheduled",
        "Sections the solver couldn't place anywhere. Listed at the bottom "
        "of each mode sheet so you can see which classes need attention."),
    6: ("Must-Have Met",
        "How many high-priority (must-have) sections were successfully "
        "placed. Anything less than full is a red flag — the solver "
        "couldn't honor a class you marked critical."),
}


# ---------------------------------------------------------------------------
# Colour palette
# ---------------------------------------------------------------------------

def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


DEPT_FILL = {
    "game":         _fill("BDD7EE"),   # soft blue
    "motion_media": _fill("E2CFEA"),   # soft purple
    "ai":           _fill("C6EFCE"),   # soft green
}
AFFINITY_FILL  = _fill("FFFF99")       # yellow  — preferred / override prof
TIME_BAD_FILL  = _fill("FFB347")       # orange  — not_preferred time slot
HEADER_FILL    = _fill("2F5496")       # dark navy header
ALT_FILL       = _fill("F2F2F2")       # light grey for alternating blank rows
WHITE_FILL     = _fill("FFFFFF")

HEADER_FONT    = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
BODY_FONT      = Font(name="Calibri", size=10)
BOLD_FONT      = Font(bold=True, name="Calibri", size=10)
TITLE_FONT     = Font(bold=True, name="Calibri", size=14)

CENTER  = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT    = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_thin = Side(style="thin", color="CCCCCC")
THIN_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _set_col_widths(ws, widths: dict) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _header_row(ws, row: int, values: list[str]) -> None:
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col, value=val)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _data_cell(ws, row: int, col: int, value, fill=None, bold=False) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill      = fill or WHITE_FILL
    cell.font      = BOLD_FONT if bold else BODY_FONT
    cell.alignment = LEFT
    cell.border    = THIN_BORDER


def _days_label(dg: int) -> str:
    return "/".join(DAY_GROUPS.get(dg, [])) or f"dg{dg}"


def _section_label(catalog_id: str, section_idx: int, total_sections: int) -> str:
    if total_sections > 1:
        return f"{catalog_id} (S{section_idx + 1})"
    return catalog_id


# ---------------------------------------------------------------------------
# Summary sheet
# ---------------------------------------------------------------------------

def _write_summary(ws, results: dict) -> None:
    quarter = results["quarter"].capitalize()
    year    = results["year"]
    modes   = results["modes"]

    ws.title = "Summary"
    ws.freeze_panes = "A4"

    # Title
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"SCAD Game Dept Course Schedule  —  {quarter} {year}"
    c.font  = TITLE_FONT
    c.alignment = CENTER
    c.fill = HEADER_FILL
    c.font = Font(bold=True, name="Calibri", size=14, color="FFFFFF")

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = f"Generated: {date.today().isoformat()}"
    c.font  = BODY_FONT
    c.alignment = CENTER

    # Mode comparison
    r = 4
    ws.merge_cells(f"A{r}:F{r}")
    ws[f"A{r}"].value = "SCHEDULE OPTIONS COMPARISON"
    ws[f"A{r}"].font  = BOLD_FONT
    ws[f"A{r}"].alignment = CENTER

    r += 1
    _header_row(ws, r, ["Mode", "Status", "Penalty Score", "Placed", "Unscheduled", "Must-Have Met"])
    # Hover comments on the four metric columns — teaches what each number means
    # for readers who didn't run the solve themselves.
    for col, (_label, text) in _SUMMARY_COMMENTS.items():
        ws.cell(row=r, column=col).comment = Comment(text, _COMMENT_AUTHOR)

    for res in modes:
        r += 1
        n_placed  = len(res["schedule"])
        n_total   = n_placed + len(res["unscheduled"])
        # Derive Must-Have Met from per-entry priorities (each schedule and
        # unscheduled item carries its own priority). No dependency on the
        # solver's `data` lookup table — that field gets stripped when
        # solver_results is embedded in the hidden _state sheet.
        n_must_unmet = sum(1 for u in res["unscheduled"] if u.get("priority") == "must_have")
        n_must = sum(
            1 for e in (res["schedule"] + res["unscheduled"])
            if e.get("priority") == "must_have"
        )
        must_met  = f"{n_must - n_must_unmet}/{n_must}"

        mode_label = res["mode"].replace("_", " ").title()
        obj = res["objective"] if res["objective"] is not None else "—"
        row_fill = _fill("E8F5E9") if res["status"] in ("optimal", "feasible") else _fill("FFEBEE")

        for col, val in enumerate(
            [mode_label, res["status"].upper(), obj,
             f"{n_placed}/{n_total}", len(res["unscheduled"]), must_met],
            start=1
        ):
            cell = ws.cell(row=r, column=col, value=val)
            cell.fill      = row_fill
            cell.font      = BODY_FONT
            cell.alignment = CENTER
            cell.border    = THIN_BORDER

    # Quarter overview — only renders when a mode carries the solver's
    # course_sections lookup. After a reload-without-resolve the embedded
    # solver_results has its `data` stripped (CP-SAT artifacts don't survive
    # JSON), so this block degrades gracefully rather than crashing.
    r += 2
    any_data = next(
        (m for m in modes
         if m["schedule"]
         and isinstance(m.get("data"), dict)
         and m["data"].get("course_sections")),
        None,
    )
    if any_data:
        sections = any_data["data"]["course_sections"]
        from collections import Counter
        priority_counts = Counter(cs["offering"]["priority"] for cs in sections)
        dept_counts     = Counter(cs["course"]["department"] for cs in sections)

        ws.merge_cells(f"A{r}:F{r}")
        ws[f"A{r}"].value = "QUARTER OVERVIEW"
        ws[f"A{r}"].font  = BOLD_FONT
        ws[f"A{r}"].alignment = CENTER

        for label, val in [
            ("Quarter",      f"{quarter} {year}"),
            ("Offerings",    f"{len(set(cs['catalog_id'] for cs in sections))} courses, {len(sections)} sections"),
            ("Must-have sections",   priority_counts.get("must_have", 0)),
            ("Should-have sections", priority_counts.get("should_have", 0)),
            ("Could-have sections",  priority_counts.get("could_have", 0)),
            ("Game dept sections",        dept_counts.get("game", 0)),
            ("Motion media sections",     dept_counts.get("motion_media", 0)),
            ("AI sections",               dept_counts.get("ai", 0)),
        ]:
            r += 1
            ws.cell(row=r, column=1, value=label).font = BOLD_FONT
            ws.cell(row=r, column=2, value=str(val)).font  = BODY_FONT

    # Colour legend
    r += 2
    ws.merge_cells(f"A{r}:F{r}")
    ws[f"A{r}"].value = "COLOUR LEGEND"
    ws[f"A{r}"].font  = BOLD_FONT
    for legend_fill, label in [
        (DEPT_FILL["game"],         "Game dept course"),
        (DEPT_FILL["motion_media"], "Motion Media dept course"),
        (DEPT_FILL["ai"],           "AI dept course"),
        (AFFINITY_FILL,             "Prof is preferred/override (good affinity)"),
        (TIME_BAD_FILL,             "Time slot not preferred for this prof"),
    ]:
        r += 1
        ws.cell(row=r, column=1).fill = legend_fill
        ws.cell(row=r, column=1).border = THIN_BORDER
        ws.cell(row=r, column=2, value=label).font = BODY_FONT

    _set_col_widths(ws, {"A": 24, "B": 18, "C": 14, "D": 12, "E": 14, "F": 16})
    ws.row_dimensions[1].height = 22


# ---------------------------------------------------------------------------
# Schedule sheet (one per mode)
# ---------------------------------------------------------------------------

_SCHED_HEADERS = [
    "Time Slot", "Days", "Course ID", "Sec",
    "Course Name", "G?", "Professor", "Room",
    "Priority", "Affinity", "Time Pref",
]
_SCHED_WIDTHS = {
    "A": 12, "B": 10, "C": 14, "D": 5,
    "E": 42, "F": 5,  "G": 22, "H": 16,
    "I": 12, "J": 16, "K": 14,
}

_AFFINITY_LABELS = {0: "Override", 1: "Preferred", 2: "Eligible", 3: "Fallback"}
_PRIORITY_LABELS = {"must_have": "Must", "should_have": "Should", "could_have": "Could"}


def _write_schedule_sheet(ws, result: dict) -> None:
    mode   = result["mode"]
    status = result["status"]

    ws.title = mode.replace("_", " ").title()
    ws.freeze_panes = "A3"

    # Title row
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value = f"Mode: {mode.replace('_', ' ').title()}  |  Status: {status.upper()}  |  Score: {result['objective'] or 'N/A'}"
    c.font  = TITLE_FONT
    c.fill  = HEADER_FILL
    c.font  = Font(bold=True, name="Calibri", size=12, color="FFFFFF")
    c.alignment = CENTER

    _header_row(ws, 2, _SCHED_HEADERS)
    _set_col_widths(ws, _SCHED_WIDTHS)
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 16

    if not result["schedule"]:
        ws.cell(row=3, column=1, value="No schedule produced — see Summary sheet.").font = BOLD_FONT
        return

    # Count total sections per catalog_id (for section label suffix)
    from collections import Counter
    section_counts = Counter(a["catalog_id"] for a in result["schedule"])

    # Group by time slot for visual separation
    current_ts = None
    r = 3

    for a in result["schedule"]:
        # Blank separator between time slots
        if a["time_slot"] != current_ts and current_ts is not None:
            for col in range(1, len(_SCHED_HEADERS) + 1):
                ws.cell(row=r, column=col).fill = ALT_FILL
            r += 1
        current_ts = a["time_slot"]

        dept_fill = DEPT_FILL.get(a["department"], WHITE_FILL)
        aff_fill  = AFFINITY_FILL if a["affinity_level"] <= 1 else dept_fill
        time_fill = TIME_BAD_FILL if a["time_pref"] == "not_preferred" else dept_fill

        course_id_label = _section_label(
            a["catalog_id"], a["section_idx"], section_counts[a["catalog_id"]]
        )
        days_label  = _days_label(a["day_group"])
        grad_label  = "G" if a["is_graduate"] else ""
        prio_label  = _PRIORITY_LABELS.get(a["priority"], a["priority"])
        aff_label   = _AFFINITY_LABELS.get(a["affinity_level"], "?")
        time_label  = a["time_pref"].replace("_", " ").title()

        values = [
            a["time_slot"], days_label, course_id_label, a["section_idx"] + 1,
            a["course_name"], grad_label, a["prof_name"], a["room_name"],
            prio_label, aff_label, time_label,
        ]
        fills = [
            dept_fill, dept_fill, dept_fill, dept_fill,
            dept_fill, dept_fill, dept_fill, dept_fill,
            dept_fill, aff_fill, time_fill,
        ]

        for col, (val, fill) in enumerate(zip(values, fills), start=1):
            _data_cell(ws, r, col, val, fill=fill)
        ws.row_dimensions[r].height = 15
        r += 1

    # Unscheduled section at bottom
    if result["unscheduled"]:
        r += 1
        ws.merge_cells(f"A{r}:K{r}")
        ws[f"A{r}"].value = "UNSCHEDULED SECTIONS"
        ws[f"A{r}"].font  = BOLD_FONT
        ws[f"A{r}"].fill  = _fill("FFD7D7")
        r += 1

        for u in result["unscheduled"]:
            ws.cell(row=r, column=1, value=u["cs_key"]).font   = BODY_FONT
            ws.cell(row=r, column=2, value=u["priority"]).font = BODY_FONT
            r += 1


# ---------------------------------------------------------------------------
# Hidden data sheets — the workbook-is-SSOT side of the export
# ---------------------------------------------------------------------------

def _write_kv_sheet(ws, sheet_name: str, kv: dict) -> None:
    """Two-column key/value sheet. Column A = key, column B = value; nested
    dicts/lists are JSON-stringified so values round-trip through a cell."""
    ws.title = sheet_name
    ws.sheet_state = "veryHidden"
    for row_idx, (k, v) in enumerate(kv.items(), start=1):
        ws.cell(row=row_idx, column=1, value=k)
        if isinstance(v, (dict, list)):
            v = json.dumps(v, separators=(",", ":"), default=str)
        ws.cell(row=row_idx, column=2, value=v)


def _write_flat_table_sheet(ws, sheet_name: str, entities: list[dict]) -> None:
    """Row-per-entity table with a header row. Columns are the union of keys
    across ``entities`` in first-seen order; nested dict/list values get
    JSON-stringified so the sheet stays rectangular."""
    ws.title = sheet_name
    ws.sheet_state = "veryHidden"
    if not entities:
        return
    keys: list[str] = []
    seen: set[str] = set()
    for e in entities:
        for k in e.keys():
            if k not in seen:
                seen.add(k)
                keys.append(k)
    for col, k in enumerate(keys, start=1):
        ws.cell(row=1, column=col, value=k)
    for row_idx, e in enumerate(entities, start=2):
        for col, k in enumerate(keys, start=1):
            v = e.get(k)
            if isinstance(v, (dict, list)):
                v = json.dumps(v, separators=(",", ":"), default=str)
            ws.cell(row=row_idx, column=col, value=v)


def _write_json_chunks_sheet(ws, sheet_name: str, payload) -> None:
    """Chunk a nested JSON payload across column A. Used for solver results
    (deeply nested, rarely inspected — a flat table would be overkill)."""
    ws.title = sheet_name
    ws.sheet_state = "veryHidden"
    serialized = json.dumps(payload, separators=(",", ":"), default=str)
    if not serialized or serialized == "null":
        return
    for i, start in enumerate(range(0, len(serialized), _DATA_CHUNK_SIZE)):
        ws.cell(
            row=1 + i,
            column=1,
            value=serialized[start:start + _DATA_CHUNK_SIZE],
        )


def _write_data_sheets(wb, draft_state: dict) -> None:
    """Embed user's working state across veryHidden ``_data_*`` sheets.

    Sheets written
    --------------
      _data_meta             marker, schema_version + scalar fields
                             (exported_at, source, quarter, year, solver_mode)
      _data_offerings        row-per-offering flat table  (required)
      _data_professors       row-per-professor flat table (optional)
      _data_rooms            row-per-room flat table      (optional)
      _data_locked_assignments  Streamlit locks table     (optional)
      _data_tuned_weights    key/value scalars            (optional)
      _data_solver_results   JSON-chunked cache           (optional)

    Optional sheets are skipped when the corresponding draft_state key is
    ``None`` — the reader maps absent sheets back to absent keys.
    """
    # Meta — always-present scalars plus the internal marker / schema_version.
    meta_kv: dict = {
        "marker":         DATA_MARKER,
        "schema_version": DATA_SCHEMA_VERSION,
    }
    for k in ("exported_at", "source", "quarter", "year", "solver_mode"):
        if k in draft_state:
            meta_kv[k] = draft_state[k]
    _write_kv_sheet(wb.create_sheet(), DATA_SHEET_META, meta_kv)

    # List-valued entity sheets
    for key, sheet_name in _DATA_LIST_SHEETS.items():
        entities = draft_state.get(key)
        if entities is None:
            continue
        _write_flat_table_sheet(wb.create_sheet(), sheet_name, entities)

    # Tuned weights (optional scalar kv)
    tw = draft_state.get("tunedWeights")
    if tw is not None:
        _write_kv_sheet(wb.create_sheet(), DATA_SHEET_TUNED_WEIGHTS, tw)

    # Solver results (optional, nested, chunked)
    sr = draft_state.get("solver_results")
    if sr is not None:
        _write_json_chunks_sheet(wb.create_sheet(), DATA_SHEET_SOLVER_RESULTS, sr)


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def write_excel(
    results: dict,
    output_dir: Path,
    draft_state: dict | None = None,
    backup_root: Path | None = None,
) -> Path:
    """Write the workbook and return the file path.

    If ``draft_state`` is provided, six veryHidden ``_data_*`` sheets are
    appended carrying the user's working state so the file can be re-opened
    later as a draft. The workbook structure is also password-protected to
    keep chairs from accidentally unhiding or reorganizing those sheets.

    If ``backup_root`` is provided, the written workbook is also copied into
    ``<backup_root>/.backups/<stem>_<ISO>.xlsx`` and that directory is pruned
    to the ``BACKUP_RETENTION`` newest entries. Backup failures are logged to
    the exception, not re-raised — a disaster-recovery copy must never take
    down the export itself.
    """
    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    quarter = results["quarter"]
    year    = results["year"]
    fname   = f"schedule_{quarter}_{year}.xlsx"
    path    = output_dir / fname

    wb = openpyxl.Workbook()

    # Sheet 1: Summary (use the default sheet)
    ws_summary = wb.active
    _write_summary(ws_summary, results)

    # Sheets 2-4: one per mode
    for res in results["modes"]:
        ws = wb.create_sheet()
        _write_schedule_sheet(ws, res)

    # Hidden data sheets — only when caller supplies a draft state
    if draft_state is not None:
        _write_data_sheets(wb, draft_state)
        wb.security = WorkbookProtection(
            workbookPassword=_DATA_STRUCTURE_PASSWORD,
            lockStructure=True,
            lockWindows=False,
        )

    wb.save(path)

    if backup_root is not None:
        _save_backup(path, Path(backup_root))

    return path


def _save_backup(src: Path, backup_root: Path) -> None:
    """Copy ``src`` into ``backup_root/.backups/<stem>_<ISO>.xlsx`` and prune
    the directory to the ``BACKUP_RETENTION`` most recent *.xlsx files.

    Swallows every exception — if the disaster-recovery copy fails we'd
    rather the user still gets their export than see a cryptic error.
    """
    try:
        backups_dir = backup_root / BACKUPS_DIRNAME
        backups_dir.mkdir(exist_ok=True)
        iso = datetime.now().strftime("%Y%m%dT%H%M%S")
        dest = backups_dir / f"{src.stem}_{iso}.xlsx"
        dest.write_bytes(src.read_bytes())

        surviving = sorted(
            backups_dir.glob("*.xlsx"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        for stale in surviving[BACKUP_RETENTION:]:
            try:
                stale.unlink()
            except OSError:
                pass
    except Exception:  # noqa: BLE001 — backup failure must not break export
        pass
