"""Tests for export/excel_reader.py — the reload side of the workbook SSOT.

Round-trips against excel_writer (write a draft state, read it back,
assert equality), exercises every typed exception, and pins the
reference-drift policy (drop + warn, never refuse the file).
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from export.excel_reader import (
    MalformedState,
    MarkerMismatch,
    MissingStateSheet,
    SchemaVersionUnsupported,
    StateReadError,
    read_draft_state,
    validate_against_local_data,
)
from export.excel_writer import (
    DATA_MARKER,
    DATA_SCHEMA_VERSION,
    DATA_SHEET_META,
    DATA_SHEET_OFFERINGS,
    write_excel,
)


def _minimal_results() -> dict:
    return {
        "quarter": "fall",
        "year": 2026,
        "modes": [
            {
                "mode": "balanced",
                "status": "optimal",
                "objective": 0,
                "schedule": [],
                "unscheduled": [],
                "data": {"priority_by_cs_key": {}},
            }
        ],
    }


def _sample_state() -> dict:
    return {
        "schema_version": DATA_SCHEMA_VERSION,
        "exported_at": "2026-04-17T12:00:00",
        "quarter": "fall",
        "year": 2026,
        "offerings": [
            {"catalog_id": "ITGM-220", "priority": "must_have", "sections": 1},
            {"catalog_id": "ITGM-340", "priority": "should_have", "sections": 2},
        ],
        "locked_assignments": [
            {
                "cs_key": "ITGM-220__0",
                "prof_id": "p_001",
                "room_id": "r_101",
                "day_group": 1,
                "time_slot": "8:00 AM",
            }
        ],
        "solver_mode": "affinity_first",
    }


# ---------------------------------------------------------------------------
# Reader — happy path + exceptions
# ---------------------------------------------------------------------------

def test_round_trip_preserves_top_level_keys(tmp_path: Path) -> None:
    original = _sample_state()
    path = write_excel(_minimal_results(), tmp_path, draft_state=original)
    loaded = read_draft_state(path)
    for k in (
        "schema_version",
        "exported_at",
        "quarter",
        "year",
        "solver_mode",
        "offerings",
        "locked_assignments",
    ):
        assert loaded[k] == original[k], f"key {k} didn't round-trip"


def test_read_from_bytes_works(tmp_path: Path) -> None:
    """Streamlit's UploadedFile gives us bytes — must accept that path."""
    path = write_excel(_minimal_results(), tmp_path, draft_state=_sample_state())
    raw = path.read_bytes()
    loaded = read_draft_state(raw)
    assert loaded["quarter"] == "fall"


def test_missing_data_meta_raises_specific(tmp_path: Path) -> None:
    """Workbooks without _data_meta (e.g. older exports) raise MissingStateSheet."""
    path = write_excel(_minimal_results(), tmp_path)  # no draft_state passed
    with pytest.raises(MissingStateSheet):
        read_draft_state(path)


def test_marker_mismatch_raises_specific(tmp_path: Path) -> None:
    path = write_excel(_minimal_results(), tmp_path, draft_state=_sample_state())
    # Corrupt the marker value in _data_meta (row 1: A=key, B=value)
    wb = openpyxl.load_workbook(path)
    wb[DATA_SHEET_META]["B1"] = "NOT_OUR_MARKER"
    bad = tmp_path / "bad_marker.xlsx"
    wb.save(bad)
    with pytest.raises(MarkerMismatch):
        read_draft_state(bad)


def test_unsupported_schema_version_carries_versions(tmp_path: Path) -> None:
    # Write at supported version, then bump the cell in _data_meta (row 2, col B)
    path = write_excel(_minimal_results(), tmp_path, draft_state=_sample_state())
    wb = openpyxl.load_workbook(path)
    wb[DATA_SHEET_META]["B2"] = DATA_SCHEMA_VERSION + 99
    bad = tmp_path / "future_schema.xlsx"
    wb.save(bad)
    with pytest.raises(SchemaVersionUnsupported) as exc_info:
        read_draft_state(bad)
    assert exc_info.value.found == DATA_SCHEMA_VERSION + 99
    assert exc_info.value.supported == DATA_SCHEMA_VERSION


def test_missing_required_keys_raises_specific(tmp_path: Path) -> None:
    """A draft state missing a required meta key (solver_mode) should fail
    the required-key check on reload — even if every other meta field is fine."""
    partial_state = _sample_state()
    del partial_state["solver_mode"]  # writer skips absent keys
    path = write_excel(_minimal_results(), tmp_path, draft_state=partial_state)
    with pytest.raises(MalformedState) as exc_info:
        read_draft_state(path)
    assert "solver_mode" in str(exc_info.value)


def test_all_typed_errors_inherit_state_read_error(tmp_path: Path) -> None:
    """UI code can catch StateReadError as a single net for any failure."""
    path = write_excel(_minimal_results(), tmp_path)  # no _data_meta
    with pytest.raises(StateReadError):
        read_draft_state(path)


# ---------------------------------------------------------------------------
# Entity sheets — shape + optional-sheet semantics
# ---------------------------------------------------------------------------

def test_optional_sheets_absent_when_state_key_not_set(tmp_path: Path) -> None:
    """If draft_state has no `professors` key, the sheet must not be created —
    and the reader must not invent an empty list."""
    state = _sample_state()  # no professors / rooms / tunedWeights
    path = write_excel(_minimal_results(), tmp_path, draft_state=state)
    wb = openpyxl.load_workbook(path)
    assert "_data_professors" not in wb.sheetnames
    assert "_data_rooms" not in wb.sheetnames
    assert "_data_tuned_weights" not in wb.sheetnames

    loaded = read_draft_state(path)
    assert "professors" not in loaded
    assert "rooms" not in loaded
    assert "tunedWeights" not in loaded


def test_professors_and_rooms_roundtrip_as_flat_tables(tmp_path: Path) -> None:
    state = _sample_state()
    state["professors"] = [
        {"id": "p_001", "name": "Smith", "department": "game"},
        {"id": "p_002", "name": "Jones", "department": "ai", "preferred_times": ["8:00 AM", "11:00 AM"]},
    ]
    state["rooms"] = [
        {"id": "r_101", "name": "Lab 101", "capacity": 30},
    ]
    path = write_excel(_minimal_results(), tmp_path, draft_state=state)
    loaded = read_draft_state(path)
    assert loaded["professors"] == state["professors"]
    assert loaded["rooms"] == state["rooms"]


def test_tuned_weights_roundtrip(tmp_path: Path) -> None:
    state = _sample_state()
    state["tunedWeights"] = {"affinity": 50, "time_pref": 30, "overload": 20}
    path = write_excel(_minimal_results(), tmp_path, draft_state=state)
    loaded = read_draft_state(path)
    assert loaded["tunedWeights"] == state["tunedWeights"]


def test_nested_offering_fields_survive_roundtrip(tmp_path: Path) -> None:
    """Offerings have nested `pinned` and list `override_preferred_professors`
    fields — both must survive the flat-table JSON-stringify/parse cycle."""
    state = _sample_state()
    state["offerings"] = [
        {
            "catalog_id": "GAME-300",
            "priority": "must_have",
            "sections": 1,
            "pinned": {"day_group": 2, "time_slot": "2:00 PM"},
            "override_preferred_professors": ["p_001", "p_002"],
        }
    ]
    path = write_excel(_minimal_results(), tmp_path, draft_state=state)
    loaded = read_draft_state(path)
    assert loaded["offerings"] == state["offerings"]


# ---------------------------------------------------------------------------
# validate_against_local_data — drift policy
# ---------------------------------------------------------------------------

def test_validate_passes_through_when_all_refs_resolve() -> None:
    state = _sample_state()
    cleaned, warnings = validate_against_local_data(
        state,
        catalog_ids={"ITGM-220", "ITGM-340"},
        prof_ids={"p_001"},
        room_ids={"r_101"},
    )
    assert warnings == []
    assert cleaned["offerings"] == state["offerings"]
    assert cleaned["locked_assignments"] == state["locked_assignments"]


def test_validate_drops_offering_with_unknown_catalog_id() -> None:
    state = _sample_state()
    cleaned, warnings = validate_against_local_data(
        state,
        catalog_ids={"ITGM-220"},  # missing ITGM-340
        prof_ids={"p_001"},
        room_ids={"r_101"},
    )
    assert len(cleaned["offerings"]) == 1
    assert cleaned["offerings"][0]["catalog_id"] == "ITGM-220"
    assert len(warnings) == 1
    assert "ITGM-340" in warnings[0]
    assert "1 of 2 offerings" in warnings[0]


def test_validate_drops_lock_with_unknown_prof() -> None:
    state = _sample_state()
    cleaned, warnings = validate_against_local_data(
        state,
        catalog_ids={"ITGM-220", "ITGM-340"},
        prof_ids=set(),  # p_001 not present
        room_ids={"r_101"},
    )
    assert cleaned["locked_assignments"] == []
    assert len(warnings) == 1
    assert "p_001" in warnings[0]
    assert "professor" in warnings[0]


def test_validate_drops_lock_with_unknown_room() -> None:
    state = _sample_state()
    cleaned, warnings = validate_against_local_data(
        state,
        catalog_ids={"ITGM-220", "ITGM-340"},
        prof_ids={"p_001"},
        room_ids=set(),  # r_101 not present
    )
    assert cleaned["locked_assignments"] == []
    assert len(warnings) == 1
    assert "r_101" in warnings[0]
    assert "room" in warnings[0]


def test_validate_collapses_long_drop_lists() -> None:
    """Don't spam the user — sample first few, summarize the rest."""
    state = _sample_state()
    state["offerings"] = [
        {"catalog_id": f"GHOST-{i:03d}", "priority": "could_have", "sections": 1}
        for i in range(20)
    ]
    cleaned, warnings = validate_against_local_data(
        state,
        catalog_ids={"ITGM-220"},  # none of the GHOSTs match
        prof_ids={"p_001"},
        room_ids={"r_101"},
    )
    assert cleaned["offerings"] == []
    assert len(warnings) == 1
    assert "20" in warnings[0]
    assert "+15 more" in warnings[0]  # 20 dropped, sample 5, +15 more


def test_validate_preserves_all_other_top_level_keys() -> None:
    state = _sample_state()
    state["custom_extension"] = {"future": "field"}
    cleaned, _ = validate_against_local_data(
        state,
        catalog_ids={"ITGM-220", "ITGM-340"},
        prof_ids={"p_001"},
        room_ids={"r_101"},
    )
    assert cleaned["custom_extension"] == {"future": "field"}
    assert cleaned["solver_mode"] == "affinity_first"
    assert cleaned["schema_version"] == DATA_SCHEMA_VERSION


# ---------------------------------------------------------------------------
# Embedded solver_results (the empty-calendar-on-reload fix)
# ---------------------------------------------------------------------------

def test_round_trip_preserves_embedded_solver_results(tmp_path: Path) -> None:
    """When the writer embeds the last solver output, the reader must return
    it intact — that's how the calendar repopulates on reload without a
    forced re-solve."""
    state = _sample_state()
    state["solver_results"] = {
        "quarter": "fall",
        "year": 2026,
        "modes": [
            {
                "mode": "balanced",
                "status": "optimal",
                "objective": 42,
                "schedule": [
                    {
                        "cs_key": "ITGM-220__0",
                        "catalog_id": "ITGM-220",
                        "prof_id": "p_001",
                        "room_id": "r_101",
                        "day_group": 1,
                        "time_slot": "8:00 AM",
                        "course_name": "Game Programming",
                        "section_idx": 0,
                        "department": "game",
                        "is_graduate": False,
                        "priority": "must_have",
                        "affinity_level": 0,
                        "time_pref": "preferred",
                        "prof_name": "Smith",
                        "room_name": "Lab 101",
                    }
                ],
                "unscheduled": [],
                "data": {"priority_by_cs_key": {"ITGM-220__0": "must_have"}},
            }
        ],
    }
    path = write_excel(_minimal_results(), tmp_path, draft_state=state)
    loaded = read_draft_state(path)
    assert loaded["solver_results"] == state["solver_results"]


def test_validate_passes_solver_results_through_unchanged() -> None:
    """validate_against_local_data only filters offerings + locks. The
    solver_results key flows through verbatim — orphan-prevention is the
    hydrate step's job (it drops solver_results when warnings are present)."""
    state = _sample_state()
    state["solver_results"] = {"modes": [{"mode": "balanced", "schedule": []}]}
    cleaned, warnings = validate_against_local_data(
        state,
        catalog_ids={"ITGM-220", "ITGM-340"},
        prof_ids={"p_001"},
        room_ids={"r_101"},
    )
    assert warnings == []
    assert cleaned["solver_results"] == state["solver_results"]
