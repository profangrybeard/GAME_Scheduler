"""Tests for the React-side roundtrip endpoints:

  POST /api/export        — must embed _data_* sheets
  POST /api/state/parse   — must read them back, validate against local data,
                            map each typed reader exception to a specific
                            HTTP status with user-facing detail copy

These run without a live solver (`run_schedule` is monkeypatched to return
a canned result) so they stay in CI's fast lane.
"""
from __future__ import annotations

import io
import json
from pathlib import Path
from typing import Callable

import openpyxl
import pytest
from fastapi.testclient import TestClient

from export.excel_reader import read_draft_state
from export.excel_writer import (
    DATA_MARKER,
    DATA_SCHEMA_VERSION,
    DATA_SHEET_META,
    write_excel,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def client() -> TestClient:
    from api.server import app
    return TestClient(app)


def _minimal_export_body(*, solve_mode: str = "balanced") -> dict:
    """A request body the export endpoint will accept. Real catalog IDs so
    react_offerings_to_doc → run_schedule fake plays nicely with the rest of
    the adapter pipeline."""
    return {
        "quarter": "fall",
        "year": 2026,
        "solveMode": solve_mode,
        "offerings": [
            {
                "catalog_id": "GAME_120",
                "priority": "must_have",
                "sections": 1,
            },
            {
                "catalog_id": "GAME_121",
                "priority": "should_have",
                "sections": 1,
                "pinned": {"day_group": 1, "time_slot": "8:00 AM"},
            },
        ],
        "professors": [],
        "rooms": [],
    }


def _fake_solver_result(quarter: str = "fall") -> dict:
    """Synthetic results dict matching the shape `run_schedule` returns.
    The `data` field carries a tuple-keyed dict that JSON would refuse —
    that's intentional, so the test can assert _strip_unserializable_results
    actually strips it before embedding."""
    return {
        "quarter": quarter,
        "year": 2026,
        "modes": [
            {
                "mode": "balanced",
                "status": "optimal",
                "objective": 42,
                "schedule": [],
                "unscheduled": [],
                "data": {
                    "course_sections": [],
                    "priority_by_cs_key": {},
                    # Tuple keys — the smoking gun. If any code path JSON-dumps
                    # the unstripped `data`, this raises TypeError immediately.
                    "vars_by_cs_dg_ts": {("X__0", 1, "8:00 AM"): "<cp_sat_var>"},
                },
            }
        ],
    }


@pytest.fixture
def fake_run_schedule(monkeypatch):
    """Replace solver.scheduler.run_schedule with a deterministic stub.
    Tests can rebind via monkeypatch to vary the return value."""
    def _fake(quarter, **kwargs):
        return _fake_solver_result(quarter)
    monkeypatch.setattr("solver.scheduler.run_schedule", _fake)
    return _fake


# ---------------------------------------------------------------------------
# /api/export — embeds the _data_* sheets
# ---------------------------------------------------------------------------

def test_export_returns_xlsx_response(client, fake_run_schedule):
    res = client.post("/api/export", json=_minimal_export_body())
    assert res.status_code == 200
    assert res.headers["content-type"] == \
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    assert "attachment" in res.headers["content-disposition"]


def test_export_response_has_very_hidden_meta_sheet(client, fake_run_schedule):
    res = client.post("/api/export", json=_minimal_export_body())
    wb = openpyxl.load_workbook(io.BytesIO(res.content))
    assert DATA_SHEET_META in wb.sheetnames
    ws = wb[DATA_SHEET_META]
    assert ws.sheet_state == "veryHidden"
    # Marker lives in cell B1 (row 1: A=key "marker", B=value DATA_MARKER)
    assert ws["B1"].value == DATA_MARKER


def test_export_embedded_state_carries_react_shape_fields(client, fake_run_schedule):
    res = client.post("/api/export", json=_minimal_export_body(solve_mode="affinity_first"))
    state = read_draft_state(res.content)

    assert state["schema_version"] == DATA_SCHEMA_VERSION
    assert state["source"] == "react"
    assert state["quarter"] == "fall"
    assert state["year"] == 2026
    assert state["solver_mode"] == "affinity_first"

    # Offerings come back in React shape, not solver-doc shape.
    assert len(state["offerings"]) == 2
    assert state["offerings"][0]["catalog_id"] == "GAME_120"
    # The pinned slot survives the round-trip
    assert state["offerings"][1]["pinned"] == {"day_group": 1, "time_slot": "8:00 AM"}


def test_export_embeds_react_shape_solver_results(client, fake_run_schedule):
    """Embedded modes must be in React shape (`assignments`, no `schedule`/`data`).
    The frontend reload reducer iterates `mode.assignments` directly — if the
    embedded modes carried the solver-native `schedule` field instead, reload
    would throw `assignments is not iterable`. Same shape as /api/solve/stream
    so React reads one shape everywhere.

    Also serves as the strip test: solver_result_to_react_mode drops `data`
    (CP-SAT decision vars + tuple-keyed indexes that aren't JSON-encodable);
    reaching this assertion at all proves the strip happened."""
    res = client.post("/api/export", json=_minimal_export_body())
    state = read_draft_state(res.content)
    embedded_results = state["solver_results"]
    assert embedded_results is not None
    for mode in embedded_results["modes"]:
        assert "assignments" in mode, \
            "embedded modes must use React shape (`assignments`, not `schedule`)"
        assert "schedule" not in mode, \
            "raw solver `schedule` field must not leak into embedded modes"
        assert "data" not in mode, \
            "per-mode `data` must be stripped (CP-SAT artifacts)"


def test_export_400_on_bad_quarter(client, monkeypatch):
    """run_schedule raises ValueError on invalid quarter — endpoint maps to 400."""
    def _raises(*a, **k):
        raise ValueError("Invalid quarter 'autumn'")
    monkeypatch.setattr("solver.scheduler.run_schedule", _raises)
    body = _minimal_export_body()
    body["quarter"] = "autumn"
    res = client.post("/api/export", json=body)
    assert res.status_code == 400
    assert "Invalid quarter" in res.json()["detail"]


# ---------------------------------------------------------------------------
# /api/state/parse — reads what /api/export wrote
# ---------------------------------------------------------------------------

def _post_parse(client: TestClient, content: bytes, filename: str = "test.xlsx"):
    return client.post(
        "/api/state/parse",
        files={"file": (filename, content,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )


def test_parse_round_trips_an_export(client, fake_run_schedule):
    """The integration test: hit /api/export, then POST that file's bytes
    back to /api/state/parse, expect the embedded state back intact."""
    exp = client.post("/api/export", json=_minimal_export_body(solve_mode="time_pref_first"))
    assert exp.status_code == 200

    res = _post_parse(client, exp.content, "schedule_fall_2026.xlsx")
    assert res.status_code == 200
    body = res.json()
    assert "state" in body
    assert "errors" in body
    assert body["errors"] == []  # clean round-trip, no drift
    assert body["state"]["solver_mode"] == "time_pref_first"
    assert body["state"]["source"] == "react"
    assert len(body["state"]["offerings"]) == 2


def test_parse_400_for_non_xlsx(client):
    """Non-zip uploads should fail fast with a friendly 400."""
    res = _post_parse(client, b"this is not an xlsx file", "garbage.xlsx")
    assert res.status_code == 400
    assert "Excel" in res.json()["detail"]


def test_parse_422_when_xlsx_has_no_data_meta(client, tmp_path):
    """Older exports / non-Scheduler XLSX files have no _data_meta sheet."""
    fake_results = {
        "quarter": "fall",
        "year": 2026,
        "modes": [{
            "mode": "balanced", "status": "optimal", "objective": 0,
            "schedule": [], "unscheduled": [],
            "data": {"priority_by_cs_key": {}},
        }],
    }
    path = write_excel(fake_results, tmp_path)  # no draft_state passed
    res = _post_parse(client, path.read_bytes())
    assert res.status_code == 422
    # Error copy mentions this is an older/non-scheduler export
    assert "older" in res.json()["detail"].lower() or "scheduler" in res.json()["detail"].lower()


def test_parse_422_with_versions_when_schema_too_new(client, tmp_path):
    """SchemaVersionUnsupported message must include both versions."""
    fake_results = {
        "quarter": "fall", "year": 2026,
        "modes": [{
            "mode": "balanced", "status": "optimal", "objective": 0,
            "schedule": [], "unscheduled": [],
            "data": {"priority_by_cs_key": {}},
        }],
    }
    state = {
        "schema_version": DATA_SCHEMA_VERSION,
        "source": "react", "quarter": "fall", "year": 2026,
        "solver_mode": "balanced",
        "offerings": [], "locked_assignments": [],
    }
    path = write_excel(fake_results, tmp_path, draft_state=state)
    # Rewrite the schema_version cell in _data_meta to a future number
    wb = openpyxl.load_workbook(path)
    wb[DATA_SHEET_META]["B2"] = DATA_SCHEMA_VERSION + 99
    future = tmp_path / "future.xlsx"
    wb.save(future)
    res = _post_parse(client, future.read_bytes())
    assert res.status_code == 422
    detail = res.json()["detail"]
    assert f"v{DATA_SCHEMA_VERSION + 99}" in detail
    assert f"v{DATA_SCHEMA_VERSION}" in detail


def test_parse_422_when_marker_corrupted(client, tmp_path):
    fake_results = {
        "quarter": "fall", "year": 2026,
        "modes": [{
            "mode": "balanced", "status": "optimal", "objective": 0,
            "schedule": [], "unscheduled": [],
            "data": {"priority_by_cs_key": {}},
        }],
    }
    valid_state = {
        "schema_version": DATA_SCHEMA_VERSION,
        "source": "react", "quarter": "fall", "year": 2026,
        "solver_mode": "balanced",
        "offerings": [], "locked_assignments": [],
    }
    path = write_excel(fake_results, tmp_path, draft_state=valid_state)
    wb = openpyxl.load_workbook(path)
    wb[DATA_SHEET_META]["B1"] = "NOT_OUR_MARKER"
    bad = tmp_path / "bad_marker.xlsx"
    wb.save(bad)
    res = _post_parse(client, bad.read_bytes())
    assert res.status_code == 422
    assert "GAME Scheduler export" in res.json()["detail"]


def test_parse_returns_structured_error_for_unknown_catalog_id(client, tmp_path):
    """Drift policy: drop offerings whose catalog_id isn't in the local
    catalog, surface one structured error per dropped record (sheet / row /
    column / reason / severity) so the Data Issues panel can render them."""
    fake_results = {
        "quarter": "fall", "year": 2026,
        "modes": [{
            "mode": "balanced", "status": "optimal", "objective": 0,
            "schedule": [], "unscheduled": [],
            "data": {"priority_by_cs_key": {}},
        }],
    }
    drift_state = {
        "schema_version": DATA_SCHEMA_VERSION,
        "source": "react", "quarter": "fall", "year": 2026,
        "solver_mode": "balanced",
        "offerings": [
            # GAME_120 exists locally; GHOST-999 doesn't
            {"catalog_id": "GAME_120", "priority": "must_have", "sections": 1},
            {"catalog_id": "GHOST-999", "priority": "could_have", "sections": 1},
        ],
        "locked_assignments": [],
    }
    path = write_excel(fake_results, tmp_path, draft_state=drift_state)
    res = _post_parse(client, path.read_bytes())
    assert res.status_code == 200
    body = res.json()
    # Only the local-known offering survives
    assert len(body["state"]["offerings"]) == 1
    assert body["state"]["offerings"][0]["catalog_id"] == "GAME_120"
    # One structured error per dropped record
    assert len(body["errors"]) == 1
    err = body["errors"][0]
    assert err["sheet"] == "_data_offerings"
    assert err["row"] == 3  # header + first offering = row 2, second offering = row 3
    assert err["column"] == "catalog_id"
    assert err["severity"] == "error"
    assert "GHOST-999" in err["reason"]
