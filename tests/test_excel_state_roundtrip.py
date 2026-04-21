"""Tests for the hidden _data_* sheets that let exported XLSX files be
re-uploaded as a working draft.

We don't run the solver here — ``write_excel`` accepts a synthetic results
dict with empty schedules so we can exercise the workbook construction +
state-roundtrip in milliseconds.
"""
from __future__ import annotations

import json
from pathlib import Path

import openpyxl
import pytest

from export.excel_writer import (
    DATA_MARKER,
    DATA_SCHEMA_VERSION,
    DATA_SHEET_META,
    DATA_SHEET_OFFERINGS,
    DATA_SHEET_PROFESSORS,
    DATA_SHEET_ROOMS,
    DATA_SHEET_SOLVER_RESULTS,
    DATA_SHEET_TUNED_WEIGHTS,
    write_excel,
)


def _minimal_results() -> dict:
    """Smallest results dict that satisfies _write_summary + _write_schedule_sheet."""
    return {
        "quarter": "fall",
        "year": 2026,
        "modes": [
            {
                "mode": "balanced",
                "status": "optimal",
                "objective": 0,
                "schedule": [],
                "unscheduled": [],
                "data": {"priority_by_cs_key": {}},
            }
        ],
    }


def _sample_draft_state() -> dict:
    return {
        "schema_version": DATA_SCHEMA_VERSION,
        "exported_at": "2026-04-17T12:00:00",
        "quarter": "fall",
        "year": 2026,
        "offerings": [
            {
                "catalog_id": "ITGM-220",
                "priority": "must_have",
                "sections": 1,
                "pinned": {"day_group": 1, "time_slot": "8:00 AM"},
            }
        ],
        "locked_assignments": [
            {
                "cs_key": "ITGM-220__0",
                "prof_id": "p_001",
                "room_id": "r_101",
                "day_group": 1,
                "time_slot": "8:00 AM",
            }
        ],
        "solver_mode": "affinity_first",
    }


def test_write_excel_without_draft_state_omits_data_sheets(tmp_path: Path) -> None:
    """Callers that don't pass draft_state get a plain 4-sheet workbook."""
    path = write_excel(_minimal_results(), tmp_path)
    wb = openpyxl.load_workbook(path)
    assert DATA_SHEET_META not in wb.sheetnames
    assert DATA_SHEET_OFFERINGS not in wb.sheetnames


def test_data_sheets_are_very_hidden_and_marker_set(tmp_path: Path) -> None:
    path = write_excel(_minimal_results(), tmp_path, draft_state=_sample_draft_state())
    wb = openpyxl.load_workbook(path)
    # Meta sheet exists, is veryHidden, and carries the marker as a kv row
    assert DATA_SHEET_META in wb.sheetnames
    ws_meta = wb[DATA_SHEET_META]
    assert ws_meta.sheet_state == "veryHidden"

    kv = {row[0].value: row[1].value for row in ws_meta.iter_rows(max_col=2) if row[0].value}
    assert kv["marker"] == DATA_MARKER
    assert kv["schema_version"] == DATA_SCHEMA_VERSION

    # Offerings sheet is also veryHidden
    assert DATA_SHEET_OFFERINGS in wb.sheetnames
    assert wb[DATA_SHEET_OFFERINGS].sheet_state == "veryHidden"


def test_workbook_structure_is_protected(tmp_path: Path) -> None:
    """When draft state is embedded, the workbook structure is locked so
    chairs can't unhide _data_* sheets via Excel's UI."""
    path = write_excel(_minimal_results(), tmp_path, draft_state=_sample_draft_state())
    wb = openpyxl.load_workbook(path)
    assert wb.security is not None
    # openpyxl exposes lockStructure as a string "1"/"0" or a bool depending on version
    assert str(wb.security.lockStructure).lower() in ("1", "true")


def test_offerings_sheet_is_a_flat_table(tmp_path: Path) -> None:
    state = _sample_draft_state()
    state["offerings"] = [
        {"catalog_id": "A", "priority": "must_have", "sections": 1},
        {"catalog_id": "B", "priority": "should_have", "sections": 2},
    ]
    path = write_excel(_minimal_results(), tmp_path, draft_state=state)
    wb = openpyxl.load_workbook(path)
    ws = wb[DATA_SHEET_OFFERINGS]

    rows = list(ws.iter_rows(values_only=True))
    headers = list(rows[0])
    assert "catalog_id" in headers
    assert "priority" in headers
    assert "sections" in headers
    # Two entity rows, header + 2 = 3
    non_empty = [r for r in rows if any(v is not None for v in r)]
    assert len(non_empty) == 3


def test_data_sheets_roundtrip_draft_state(tmp_path: Path) -> None:
    """Write a draft state, read it back via the reader, expect equality."""
    from export.excel_reader import read_draft_state
    original = _sample_draft_state()
    path = write_excel(_minimal_results(), tmp_path, draft_state=original)
    loaded = read_draft_state(path)
    # schema_version, quarter, year, solver_mode, exported_at, offerings, locks
    assert loaded["schema_version"] == DATA_SCHEMA_VERSION
    assert loaded["quarter"] == "fall"
    assert loaded["year"] == 2026
    assert loaded["solver_mode"] == "affinity_first"
    assert loaded["exported_at"] == "2026-04-17T12:00:00"
    assert loaded["offerings"] == original["offerings"]
    assert loaded["locked_assignments"] == original["locked_assignments"]


def test_solver_results_chunked_cleanly_when_large(tmp_path: Path) -> None:
    """A solver_results payload big enough to exceed Excel's 32,767-char
    single-cell limit must still roundtrip via chunking."""
    state = _sample_draft_state()
    state["solver_results"] = {
        "modes": [
            {
                "mode": "balanced",
                "assignments": [
                    {
                        "catalog_id": f"DUMMY-{i:04d}",
                        "section_idx": 0,
                        "prof_id": "p_001",
                        "room_id": "r_101",
                        "day_group": 1,
                        "time_slot": "8:00 AM",
                        "notes": "x" * 100,
                    }
                    for i in range(500)
                ],
            }
        ],
    }
    path = write_excel(_minimal_results(), tmp_path, draft_state=state)
    wb = openpyxl.load_workbook(path)
    ws = wb[DATA_SHEET_SOLVER_RESULTS]

    chunks = []
    row = 1
    while ws.cell(row=row, column=1).value:
        chunks.append(ws.cell(row=row, column=1).value)
        row += 1
    assert len(chunks) >= 2, f"expected chunking, got {len(chunks)} chunk(s)"
    parsed = json.loads("".join(chunks))
    assert parsed == state["solver_results"]


def test_summary_sheet_has_metric_column_comments(tmp_path: Path) -> None:
    """The teaching-moment hover comments on Penalty Score / Placed /
    Unscheduled / Must-Have Met must survive the round-trip through openpyxl."""
    path = write_excel(_minimal_results(), tmp_path, draft_state=_sample_draft_state())
    wb = openpyxl.load_workbook(path)
    ws = wb["Summary"]

    # Mode comparison header lives at row 5 (title=1, subtitle=2, blank=3,
    # section header=4, column headers=5). Cols 3-6 are the metric columns.
    for col in (3, 4, 5, 6):
        cell = ws.cell(row=5, column=col)
        assert cell.comment is not None, f"col {col} ({cell.value}) missing comment"
        assert len(cell.comment.text) > 30, f"col {col} comment looks too short"


def _full_assignment(cs_key: str, catalog_id: str, priority: str) -> dict:
    """All the fields _write_schedule_sheet expects per assignment row."""
    return {
        "cs_key": cs_key, "catalog_id": catalog_id, "section_idx": 0,
        "course_name": catalog_id, "department": "game", "is_graduate": False,
        "priority": priority, "affinity_level": 1, "time_pref": "preferred",
        "prof_id": "p_001", "prof_name": "Smith",
        "room_id": "r_101", "room_name": "Lab 101",
        "day_group": 1, "time_slot": "8:00 AM",
    }


def test_summary_must_have_met_works_without_data_field(tmp_path: Path) -> None:
    """After a reload-without-resolve, mode dicts have no `data` field
    (CP-SAT artifacts were stripped on export). The Must-Have Met column
    must still compute from per-entry priority, not crash with KeyError."""
    results_no_data = {
        "quarter": "fall",
        "year": 2026,
        "modes": [
            {
                "mode": "balanced",
                "status": "optimal",
                "objective": 0,
                "schedule": [
                    _full_assignment("A__0", "ITGM-A", "must_have"),
                    _full_assignment("B__0", "ITGM-B", "should_have"),
                ],
                "unscheduled": [
                    {"cs_key": "C__0", "priority": "must_have"},
                ],
                # No "data" key at all — exactly what survives a reload roundtrip
            }
        ],
    }
    path = write_excel(results_no_data, tmp_path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Summary"]
    # Mode comparison row 6, col 6 = Must-Have Met.
    # 2 must-haves total (1 placed, 1 unscheduled) → "1/2".
    assert ws.cell(row=6, column=6).value == "1/2"


def test_summary_quarter_overview_skipped_when_data_missing(tmp_path: Path) -> None:
    """Quarter Overview block uses solver `data["course_sections"]`. After
    reload-without-resolve that's gone — block must skip silently rather
    than crash. Other summary sections (mode comparison, legend) still render."""
    results_no_data = {
        "quarter": "fall",
        "year": 2026,
        "modes": [
            {
                "mode": "balanced",
                "status": "optimal",
                "objective": 0,
                "schedule": [_full_assignment("A__0", "ITGM-A", "must_have")],
                "unscheduled": [],
            }
        ],
    }
    path = write_excel(results_no_data, tmp_path)
    wb = openpyxl.load_workbook(path)
    ws = wb["Summary"]
    all_text = " ".join(
        str(c.value) for row in ws.iter_rows() for c in row if c.value
    )
    assert "QUARTER OVERVIEW" not in all_text
    assert "COLOUR LEGEND" in all_text
